from datetime import datetime
from openpyxl import load_workbook

rut = "C:\Users\SENA\Desktop\Ruta 1\Base Crud.xlsx"

def leer (ruta: str, extraer: str):
    Archivo_Exccel = load_workbook(ruta)
    Hoja_Datos = Archivo_Exccel ["Datos del Crud"]
    Hoja_Datos = Hoja_Datos["A2": "F" + str(Hoja_Datos.max_row)]

    info = {}

    for i in Hoja_Datos:
        if isinstance(i[0].value, int):
            info.setdefault(i[0].value, {"Tarea": i[1].value, "Descripcion": i[2].value, "Estado": i[3].value, "Fecha de Inicio":i[4].value, "Fecha de Finalizacion": i[5].value})
    
    if not(extraer == "todo"):
        info = filtrar(info, extraer)
    
    for i in info:
        print("********Tarea********")
        print("Id:" + str(i) + "\n" + "Titulo: " + str(info[i]["tarea"]) + "\n" + "Descripcion: " + str(info[i]["descripcion"]) + "\n" + "Estado: " + str(info[i]["estado"]) + "\n" + "Fecha de Inicio: " + str(info[i]["fecha de inicio"]) + "\n" + "Fecha de Finalizacion: " + str(info[i]["fecha de finalizacion"]))
        print()
        
    return

def filtrar(info: dict, filtro:str):
    for i in info:
        if info[i]["estado"] == filtro:
            aux.setdefault(i, info[i])

def actualizar(ruta:str, identificador:int, datos_actualizados:dict):
    Archivo_Exccel = load_workbook(ruta)
    Hoja_Datos = Archivo_Exccel["Datos del Crud"]
    Hoja_Datos = Hoja_Datos["A2":"F" + str(Hoja_Datos.max_row)]
    
    Titulo = 2
    Descripcion = 3
    Estado = 4
    Fecha_Inicio = 5
    Fecha_Finalizacion = 6
    Encontro = False
    for i in Hoja_Datos:
        if i[0].value == identificador:
            fila = i[0].row
            Encontro = True
            for d in datos_actualizados:
                if d == "Titulo" and not(datosActualizados[d] == ""):
                    hoja.cell(row = fila, column = Titulo).value = datosActualizados[d]
                    
                elif d == "Descripcion" and not(datosActualizados[d] == ""):
                    hoja.cell(row = fila, column = Descripcion).value = datosActualizados[d]
                    
                elif d == "Estado" and not(datosActualizados[d] == ""):
                    hoja.cell(row = fila, column = Estado).value = datosActualizados[d]
                    
                elif d == "Fecha de Inicio" and not(datosActualizados[d] == ""):
                    hoja.cell(row = fila, column = Fecha_Inicio).value = datosActualizados[d]
                    
                elif d == "Fecha de Finalizacion" and not(datosActualizados[d] == ""):
                    hoja.cell(row = fila, column = Fecha_Finalizacion).value = datosActualizados[d]
                    
    Archivo_Exccel.save(ruta)
    if Encontro == False:
        print("Error: No existeuna tarea con ese Id")     
        print()
    return


def agregar(ruta:int, datos:dict):
    Archivo_Execcel = load-workbook(ruta)
    Hoja_datos = Archivo_Execcel['Datos del crud']
    Hoja_datos=Hoja_datos['A2':'F'+str(Hoja_datos.max_row+1)]
    hoja=Archivo_Execcel.active

    titulo=2
    descrripcion=3
    estado=4
    fecha_inicio=5
    fecha_finalizado=6
    for i in Hoja_datos:

        if not( isinstance(i[0].value, int)):
            identificador=i[0].row
            hoja.cell(row=identificador, column=1).value=identificador-1
            hoja.cell(row=identificador, column=titulo).value=datos['titulo']
            hoja.cell(row=identificador, column=descripcion).value=datos['descripcion']
            hoja.cell(row=identificador, column=estado).value=datos['estado']
            hoja.cell(row=identificador, column=fecha_inicio).value=datos['fecha inico']
            hoja.cell(row=identificador, column=fecha_finalizado).value=datos['fecha finalizacion']

            def borrar(ruta,identificador):
                Archivo_Execcel = load_workbook(ruta)
                Hoja_datos = Archivo_Execcel['Datos del crud']
                Hoja_datos=Hoja_datos['A2':'F'+str(Hoja_datos.max_row)]
                hoja=Archivo_Execcel.active

                titulo=2
                descripcion=3
                estado=4
                fecha_inicio=5
                fecha_finalizado=6
                encontro=False
                for i in Hoja_datos:
                    if i[0].value==identificador:
                        fila=i[0].row
                        encontro=True

                        hoja.cell(row=fila, column=1).value=""
                        hoja.cell(row=fila, column=titulo).value=""
                        hoja.cell(row=fila, column=descricion).value=""
                        hoja.cell(row=fila, column=estado).value=""
                        hoja.cell(row=fila, column=fecha_inicio).value=""
                        hoja.cell(row=fila, column=fecha_finalizado).value=""
                Archivo_Execcel.save(ruta)
                if encontro==false:
                    print('error; No existe una tarea con ese id')
                    print()
                    return

rut="C:\\Users\\Usuarioprueba\\Desktop\\Ruta1\\Base crud.xlsx"
datosActualizados={'titulo': '', 'descripcion': '', 'estado': '', 'fecha inicio': '', 'fecha finalizacion':''}

while True:
    print('Indique la accion que desea realizar: ')
    print('Consultar: 1')
    print('Actualizar: 2')
    print('Crear nueva tarea: 3')
    print('Borrar: 4')
    accion = input('Escriba la opcion:')
    
    if not (accion == '1') and not (accion=='2') and not (accion == '3') and not (accion=='4'): 
       print('Comando invalido por favor eliga una opcion valida')
    elif accion=='1':
         opc_consulta=''
         print('Indique la tarea que desea consultar: ')
         print('Todas las tareas: 1')
         print('En espera: 2')
         print('En ejecucion: 3')
         print('Por aprobar: 4')
         print('Finalizada: 5')
         opc_consulta = input('Escriba la tarea que dese consultar:')
         if opc_consulta=='1':
             print()
             print()
             print('** Consultando todas las tareas **')  
             leer (rut, 'todo')
         elif opc_consulta=='2':
             print()
             print()
             print('** Consultando tareas en espera **')
             leer (rut, 'En espera')
         elif opc_consulta=='3':
             print()
             print()
             print('** Consultando tareas en ejecucion **')
             leer (rut, 'En ejecucion')
         elif opc_consulta=='4':
             print()
             print()
             print('** Consultando tareas por aprobar **') 
             leer(rut, 'Por aprobar')
         elif opc_consulta=='5':
             print()
             print()
             print('** Consultando tareas finalizadas **')
             leer (rut,'Finalizada') 

    elif accion=='2':
        datosActualizados={'titulo': '', 'descripcion': '', 'estado': '', 'fecha inicio':'', 'fecha finalizacion':''} 
        print('** Actualizar Tarea **')
        print()
        id_Actualizar=int(input('Indique el Id de la tarea que desea actualizar: '))
        print()
        print('** Nuevo titulo **')
        print('**Nota: si no desea actualizar el titulo solo oprima ENTER') 
        datosActualizados [ 'titulo ']=input('Indique el nuevo titulo de la tarea : ')
        print()
        print('**Nota: si no desea actualizar la descripcion solo oprima ENTER')
        print('** Nueva descripcion **') 
        datosActualizados['descripcion']=input('Indique la nueva descripcion de la tarea : ')
        print()
        print('** Nueva estado **')
        print('En espera: 2')
        print('En ejecucion: 3')
        print('Por aprobar 4')
        print('Finalizada: 5')
        
        print('**Nota: si no desea actualizar el estado solo oprima ENTER') 
        estadoNuevo= input('Indique el nuevo estado de la tarea : ')
        if estadoNuevo=='2':
            datosActualizados['estado']='En espera'
        elif estadoNuevo=='3':
            datosActualizados['estado']='En ejecucion'
        elif estadoNuevo=='4':
            datosActualizados['estado']='Por aprobar'
        elif estadoNuevo=='5':
            now = datetime.now()
            datosActualizados['estado']='Finalizada'
            datosActualizados['fecha finalizacion']=str(now.day) +'/'+ str(now.month) +'/'+str(now.year)
        
        now = datetime.now()
        datosActualizados['fecha inicio']=str(now.day) +'/'+ str(now.month) +'/'+str(now.year) 
        actualizar(rut,id_Actualizar, datos Actualizados)
        print()

    elif accion=='3':
        datosActualizados={'tarea': '', 'descripcion': '', 'estado': '', 'fecha inicio': '', 'fecha finalizacion':''} 
        print('** Crear nueva Tarea **')

        print()
        print('** titulo **')
        print()
        datosActualizados ['titulo']=input('Indique el titulo de la tarea: ')
        print()
        print('** descripcion **')
        datosActualizados ['descripcion']= input('Indique la descripcion de la tarea :')
        print()
        datosActualizados ['estado']='En espera'
        now = datetime.now()
        datosActualizados ['fecha inicio']=str(now.day) +'/'+ str(now.month) +'/'+str(now.year)
        datosActualizados ['fecha finalizacion']='' 
        agregar (rut, datosActualizados)
    elif accion=='4':
        print('')
        print('*** Eliminar Tarea **')
        iden=int(input('Indique el Id de la tarea que desea eliminar: '))
        borrar (rut,iden)       